import io

from django.core.files.base import File
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN

from ..models import ExcelReport


class BaseExcelReport:

    def __init__(self, sheet_name=None):
        self.workbook = Workbook()
        self.worksheet = self.workbook.worksheets[0]
        if sheet_name:
            self.worksheet.title = sheet_name

        self._border_style = Side(border_style=BORDER_THIN, color='00000000')
        self.border = Border(
            left=self._border_style,
            right=self._border_style,
            top=self._border_style,
            bottom=self._border_style
        )

        self.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        self.now = timezone.now()

        self.filename = 'report_{}.xlsx'.format(self.now.strftime('%d.%m.%Y'))

    def _write_cell(self, cell, value, color=None):
        self.worksheet[cell] = value
        self.worksheet[cell].border = self.border
        self.worksheet[cell].alignment = self.alignment
        if color:
            self.worksheet[cell].fill = PatternFill(fgColor=color, fill_type='solid')

    def _column_width(self, column, width):
        self.worksheet.column_dimensions[column].width = width

    def _row_height(self, row, height):
        self.worksheet.row_dimensions[row].height = height

    def _merge_cells(self, cell_start, cell_end):
        self.worksheet.merge_cells(cell_start + ':' + cell_end)

    def _write_cell_hyperlink(self, cell, value, hyperlink):
        self.worksheet[cell].value = value
        self.worksheet[cell].hyperlink = hyperlink
        self.worksheet[cell].style = 'Hyperlink'
        self.worksheet[cell].border = self.border
        self.worksheet[cell].alignment = self.alignment

    def write_header(self):
        raise NotImplementedError('write_header')

    def write_data(self, qs=None):
        raise NotImplementedError('write_data')

    def for_http_response(self):
        return save_virtual_workbook(self.workbook)

    def for_django_file_field(self):
        virtual_file = io.BytesIO()
        self.workbook.save(virtual_file)
        django_file = File(virtual_file, name=self.filename)
        return django_file


class OfferExcelReport(BaseExcelReport):

    COLOR = '8ea9db'

    def _get_quarter(self):
        if 1 <= self.now.month <= 3:
            return 1

        if 4 <= self.now.month <= 6:
            return 2

        if 7 <= self.now.month <= 9:
            return 3

        if 10 <= self.now.month <= 12:
            return 4

    def write_header(self):

        self._write_cell('A1', '№ п/п')
        self._write_cell('A2', '1', color=self.COLOR)
        self._column_width('A', 10)

        self._write_cell('B1', 'Код строительного ресурса')
        self._write_cell('B2', '2', color=self.COLOR)
        self._column_width('B', 20)

        self._write_cell('C1', 'Наименование строительного ресурса')
        self._write_cell('C2', '3', color=self.COLOR)
        self._column_width('C', 20)

        self._write_cell('D1', 'Полное наименование строительного ресурса в обосновывающем документе')
        self._write_cell('D2', '4', color=self.COLOR)
        self._column_width('D', 20)

        self._write_cell('E1', 'Единицы измерения строительного ресурса')
        self._write_cell('E2', '5', color=self.COLOR)
        self._column_width('E', 20)

        self._write_cell('F1', 'Единицы измерения строительного ресурса в обосновывающем документе')
        self._write_cell('F2', '6', color=self.COLOR)
        self._column_width('F', 20)

        self._write_cell('G1', 'Текущая отпускная цена за единицу измерения в обосновывающем документе с НДС, в руб.')
        self._write_cell('G2', '7', color=self.COLOR)
        self._column_width('G', 20)

        self._write_cell('H1', 'Текущая отпускная цена за ед. изм. без НДС, в руб. в соответствии с единицей измерения строительного ресурса')
        self._write_cell('H2', '8', color=self.COLOR)
        self._column_width('H', 20)

        self._write_cell('I1', 'Цт - стоимость перевозки автомобильным транспортом на расстояние до 30 км, без НДС, руб.')
        self._write_cell('I2', '9', color=self.COLOR)
        self._column_width('I', 20)

        self._write_cell('J1', 'Цд - стоимость доставки из других субъектов РФ, без НДС, руб.')
        self._write_cell('J2', '10', color=self.COLOR)
        self._column_width('J', 20)

        self._write_cell('K1', 'Год')
        self._write_cell('K2', '11', color=self.COLOR)
        self._column_width('K', 15)

        self._write_cell('L1', 'Квартал')
        self._write_cell('L2', '12', color=self.COLOR)
        self._column_width('L', 15)

        self._write_cell('M1', 'Наименование производителя/поставщика')
        self._write_cell('M2', '13', color=self.COLOR)
        self._column_width('M', 20)

        self._write_cell('N1', 'КПП организации')
        self._write_cell('N2', '14', color=self.COLOR)
        self._column_width('N', 20)

        self._write_cell('O1', 'ИНН организации')
        self._write_cell('O2', '15', color=self.COLOR)
        self._column_width('O', 20)

        self._write_cell('P1', 'Гиперссылка на веб-сайт производителя/ поставщика')
        self._write_cell('P2', '16', color=self.COLOR)
        self._column_width('P', 20)

        self._write_cell('Q1', 'Населенный пункт расположения склада производителя/поставщика')
        self._write_cell('Q2', '17', color=self.COLOR)
        self._column_width('Q', 20)

        self._write_cell('R1', 'Статус организации (производитель - 1, поставщик - 2)')
        self._write_cell('R2', '18', color=self.COLOR)
        self._column_width('R', 20)

        self._write_cell('S1', 'Стоимость доставки до г.Оренбург')
        self._write_cell('S2', '19', color=self.COLOR)
        self._column_width('S', 20)

    def write_data(self, qs=None):
        offset = 3

        if qs is None:
            return

        for index, offer in enumerate(qs[:15]):
            self._write_cell('A' + str(offset + index), index + 1)
            offer_product = offer.product
            if offer_product:
                resource_code_str = offer_product.resource_code
                name_str = offer_product.name
            else:
                resource_code_str = offer.resource_code or ''
                name_str = offer.name
            self._write_cell('B' + str(offset + index), resource_code_str)
            self._write_cell('C' + str(offset + index), name_str)
            self._write_cell('D' + str(offset + index), offer.name)

            measure_unit_str = offer.measure_unit or ''
            self._write_cell('E' + str(offset + index), measure_unit_str)
            self._write_cell('F' + str(offset + index), measure_unit_str)

            last_price = offer.last_offer_price
            if last_price:
                price_with_vat_str = last_price.price_with_vat
                price_without_vat_str = last_price.price_without_vat
                delivery_cost_str = last_price.delivery_cost
            else:
                price_with_vat_str = ''
                price_without_vat_str = ''
                delivery_cost_str = ''
            self._write_cell('G' + str(offset + index), price_with_vat_str)
            self._write_cell('H' + str(offset + index), price_without_vat_str)
            self._write_cell('I' + str(offset + index), delivery_cost_str)
            self._write_cell('J' + str(offset + index), '')

            self._write_cell('K' + str(offset + index), self.now.strftime('%d.%m.%Y'))
            self._write_cell('L' + str(offset + index), self._get_quarter())

            self._write_cell('M' + str(offset + index), offer.provider.name)
            self._write_cell('N' + str(offset + index), offer.provider.kpp)
            self._write_cell('O' + str(offset + index), offer.provider.inn)

            if offer.page_url:
                self._write_cell_hyperlink('P' + str(offset + index), '[Открыть]', offer.page_url)
            else:
                self._write_cell('P' + str(offset + index), '')

            self._write_cell('Q' + str(offset + index), offer.provider.warehouse_location)

            self._write_cell('R' + str(offset + index), '2')
            self._write_cell('S' + str(offset + index), delivery_cost_str)

    def generate(self, qs=None):
        self.write_header()
        self.write_data(qs=qs)
        ExcelReport.objects.create(excel=self.for_django_file_field())
